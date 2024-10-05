import time

from selenium.webdriver.common.devtools.v128.pwa import change_app_user_settings

from stbk_entity import FinalPage
# from project_01 import Final_page
from stbk_entity import StarBucks
from rich.console import Console
from rich.table import Column, Table
import os

import pickle
import openpyxl as op

"""
- 커피 : 온도(Hot/Iced), 사이즈, 얼음량, 당도, 컵(텀블러 유무), 수량, 디카페인, [상품명, 샷(연하게, 보통, 진하게(샷추가+500) ), 가격 ]
- 음료 : 온도(Hot/Iced), 사이즈, 얼음량, 당도, 컵(텀블러 유무), 수량, [상품명, 휘핑 유무, 가격]
"""


def filters(result, by, value):
    filter_list = {
        '1': f'item.get_name()',
        '2': f'item.get_temp()',
        '3': f'item.get_cup()',
        '4': f'item.get_pymnt()',
        '5': f'item.get_mbr()',
        '6': f'item.get_carrier()'

    }
    os.system('cls')
    table = Table(title="==================== Total Sales List ====================", show_header=True,
                  header_style="bold magenta")
    table.add_column('No', justify='center')
    table.add_column('Menu', justify='left')
    table.add_column('Quantity', justify='center')
    table.add_column('Sales', justify='center')
    table.add_column('Accum_sales', justify='center')
    table.add_column('Togo', justify='center')
    table.add_column('Member', justify='center')  # 전화번호
    table.add_column('Payment\n1:카드 2:현금 3:기타pay', justify='center')  # '1':신용카드,'2':현금,'3':기타pay '0':NA
    table.add_column('Carrier\n1:SK 2:KT 3:LGU+ 0:NA', justify='center')  # '1':sk,'2':kt,'3':lgu '0':NA

    Accum_sales=0

    for n, item in enumerate(result):
        if eval(filter_list[by]) == value:
            Accum_sales += item.get_price()
            table.add_row(f'{n + 1}',
                          f'{item.get_size()} {item.get_temp()} {item.get_name()} / {item.get_amnt_ice()} / {item.get_sugar_cnt()}',
                          f'{item.get_quantity()}', f'{item.get_price()}', f'{Accum_sales}', f'{item.get_cup()}',
                          f'{item.get_mbr()}', f'{item.get_pymnt()}', f'{item.get_carrier()}')

    console = Console()
    console.print(table)


class StarBucks_Repository:

    order_str=[]
    sales_statement = []

    def __init__(self):
        try:
            with open('sales_statement.pkl', 'rb') as g:
                StarBucks_Repository.sales_statement = pickle.load(g)
        except:
            pass

    def __del__(self):
        with open('sales_statement.pkl', 'wb') as g:
            pickle.dump(StarBucks_Repository.sales_statement, g)




    def find_all(self):

        table = Table(title="==================== Total Sales List ====================", show_header=True, header_style="bold magenta")
        table.add_column('No',  justify='center')
        table.add_column('Menu',  justify='left')
        table.add_column('Quantity',  justify='center')
        table.add_column('Sales',  justify='center')
        table.add_column('Accum_sales',  justify='center')
        table.add_column('Togo',  justify='center')
        table.add_column('Member',  justify='center')           # 전화번호
        table.add_column('Payment\n1:카드 2:현금 3:기타pay',  justify='center')          #'1':신용카드,'2':현금,'3':기타pay '0':NA
        table.add_column('Carrier\n1:SK 2:KT 3:LGU+ 0:NA',  justify='center')  #'1':sk,'2':kt,'3':lgu '0':NA


        result = StarBucks_Repository.sales_statement
        Accum_sales = 0
        for n, item in enumerate(result):
            Accum_sales += item.get_price()
            table.add_row(f'{n+1}', f'{item.get_size()} {item.get_temp()} {item.get_name()} / {item.get_amnt_ice()} / {item.get_sugar_cnt()}',
                          f'{item.get_quantity()}',f'{item.get_price()}', f'{Accum_sales}',f'{item.get_cup()}',f'{item.get_mbr()}',f'{item.get_pymnt()}',f'{item.get_carrier()}')

        console = Console()
        console.print(table)

            # return self.stbks

        # TODO 추가했으면 하는것..커피/음료에서 얼음이 안들어가는것(HOT, 프라푸치노,...),휘핑이 원래 안들어가는것들(아메리카노, 자허블, ...)은 옵션이 안보이게 하고 싶다...

        # 1: ('녹차 프라푸치노', 6500),
        # 2: ('자바칩 프라푸치노', 7000),
        # 3: ('피치 아사이 리프레셔', 6000),
        # 4: ('딸기 아사이 리프레셔', 6000),
        # 5: ('망고 리프레셔', 6000)



    def item_sales(self):
        menu_dict={
            '1':'아메리카노',"2":'카페라떼','3':'돌체라떼','4':'카페모카','5':'캬라멜마끼아또',
            '6':'녹차라테','7':'라이트 자몽 피지오','8':'복숭아 아이스티','9':'딸기 에이드','10':'레몬 에이드',
            '11':'녹차 프라푸치노','12':'자바칩 프라푸치노','13':'피치 아사이 리프레셔','14':'딸기 아사이 리프레셔','15':'망고 리프레셔'
        }

        os.system('cls')
        table = Table(title=f"== Drink List ==", show_header=True,
                      header_style="bold magenta")
        table.add_column('No', justify='center')
        table.add_column('Menu', justify='center')

        for keys, vals in menu_dict.items():
            table.add_row(f'{keys}',f'{vals}')
        console = Console()
        console.print(table)

        menu = input('메뉴를 선택해주세요>>')

        if menu in menu_dict.keys():
            os.system('cls')
            table = Table(title=f"========== {menu_dict[menu]} Sales List ==========", show_header=True,
                          header_style="bold magenta")
            table.add_column('No', justify='center')
            table.add_column('Menu', justify='center')
            table.add_column('Accum_quantity', justify='center')
            table.add_column('Accum_sales', justify='center')

            result = StarBucks_Repository.sales_statement

            Accum_sales = 0
            Accum_quantity = 0

            for n, item in enumerate(result):
                if item.get_name() == menu_dict[menu]:
                    Accum_quantity += item.get_quantity()
                    Accum_sales += item.get_price()
                    table.add_row(f'{n + 1}', f'{item.get_name()}', f'{Accum_quantity}',f'{Accum_sales}')

            console = Console()
            console.print(table)
        else:
            print('잘못된 접근입니다.')


    def sort_by(self):
        result = StarBucks_Repository.sales_statement
        by=input('Sort by... 1:item, 2:temp, 3: togo >> ')
        match by:
            case '1':
                StarBucks_Repository.item_sales(self)
            case '2':
                value = input('1: Hot 2: Iced>>')
                filters(result, by, f'{'Hot' if value=="1" else 'Iced'}')
            case '3':
                value = input('1: Dine-in 2: To-go')
                filters(result, by, f'{'Dine-in' if value=="1" else 'To-go'}')


    def push(self, stbks):
        return self.create_workbook(stbks)
    def create_workbook(self, stbks):
        wb = op.Workbook()
        ws = wb.active
        ws['A1'].value = '번호'
        ws['B1'].value = '상품명'
        ws['C1'].value = '수량'
        ws['D1'].value = '매출액'
        # ws['E1'].value = '마진'
        # https://post.naver.com/viewer/postView.nhn?volumeNo=17083716&memberNo=6347519
        no = 1
        for stbk in stbks:
            ws[f'A{no + 1}'] = stbk.get_id()
            ws[f'B{no + 1}'] = stbk.get_name()
            ws[f'C{no + 1}'] = stbk.get_quantity()
            ws[f'D{no + 1}'] = stbk.get_sv()
            no += 1
        wb.save('starbucks.xlsx')
        print('☕☕엑셀파일이 생성되었습니다.☕☕')
        return wb

    def coffee_menu(self):
        add_on = 0
        coffee_menu_list = {
            1: ('아메리카노', 4500),
            2: ('카페라떼', 5000),
            3: ('돌체라떼', 5900),
            4: ('카페모카', 5500),
            5: ('캬라멜마끼아또', 5900)
        }
        coffee_menu_str = """
    ---------Coffee Menu--------
    1. 아메리카노 ======= 4500원
    2. 카페라떼 ========= 5000원
    3. 돌체라떼 ========= 5500원
    4. 바닐라 라떼 ====== 5500원
    5. 카라멜마키아토 ==== 5500원
    0. 이전으로 돌아가기
    ----------------------------          
    선택:  """

        os.system('cls')
        table = Table(title="========== Coffee Menu ==========",show_header=True, header_style="bold magenta")
        table.add_column('Selction',width=10, justify='center')
        table.add_column('Menu',width=30, justify='left')
        table.add_column('Price',width=10, justify='center')

        table.add_row('1','아메리카노','4500원')
        table.add_row('2','카페라떼','5000원')
        table.add_row('3','돌체라떼','5500원')
        table.add_row('4','바닐라 라떼','5500원')
        table.add_row('5','카라멜마키아토','5500원')
        table.add_row('0','이전으로 돌아가기','')


        console=Console()
        console.print(table)

        while True:
            try:
                menu_name = int(input('메뉴를 선택해 주세요.>>'))
                break
            except:
                print('Selection에 있는 "정수"로 다시 입력해주세요')

        if menu_name in coffee_menu_list:
            print(f'{coffee_menu_list[menu_name][0]}를 선택하셨습니다.')

            temp_list = ['Hot', 'Iced']

            os.system('cls')
            table = Table(title="========== Hot / Cold ==========", show_header=True, header_style="bold magenta")
            table.add_column('Selction', width=10, justify='center')
            table.add_column('Menu', width=30, justify='left')
            table.add_column('Menu', width=10, justify='center')

            table.add_row('1', 'Hot','+0원')
            table.add_row('2', 'Iced','+500원')

            console = Console()
            console.print(table)

            while True:
                try:
                    temp = int(input('원하시는 항목을 선택해 주세요 >>'))
                    if temp in [1, 2]:
                        break
                    else:
                        print('다시 입력해주세요.')
                except:
                    print('Selection에 있는 "정수"로 다시 입력해주세요')




            if temp == 2:
                add_on += 500
                ice_list = ['Extra', 'Normal', 'Less']

                os.system('cls')
                table = Table(title="========== Ice ==========", show_header=True, header_style="bold magenta")
                table.add_column('Selction', width=10, justify='center')
                table.add_column('Ice amount', width=30, justify='left')
                table.add_column('Extra Charge', width=20, justify='center')

                table.add_row('1', 'Extra', '+0원')
                table.add_row('2', 'Normal', '+0원')
                table.add_row('3', 'Less', '+0원')

                console = Console()
                console.print(table)

                while True:
                    try:
                        ice = int(input('얼음양을 선택해주세요>>'))
                        if ice in [1,2,3]:
                            break
                        else:
                            print('다시 입력해주세요.')
                    except:
                        print('Selection에 있는 "정수"로 다시 입력해주세요')


            while True:
                try:
                    qntt = int(input('\n수량을 입력해주세요.>>'))
                    break
                except:
                    print('"정수"로 다시 입력해주세요')


            swt_list = ['120%', '100%', '80%']

            os.system('cls')
            table = Table(title="========== Sweetness ==========", show_header=True, header_style="bold magenta")
            table.add_column('Selction', width=10, justify='center')
            table.add_column('Sweetness', width=30, justify='left')
            table.add_column('Extra Charge', width=10, justify='center')

            table.add_row('1', '더 달게 120%', '+300원')
            table.add_row('2', '보통 100%','+0원')
            table.add_row('3', '덜 달게 80%','+0원')

            console = Console()
            console.print(table)

            while True:
                try:
                    swt = int(input('당도를 선택해주세요>>'))
                    if swt == 1:
                        add_on += 300
                    if swt in [1, 2, 3]:
                        break
                    else:
                        print('다시 입력해주세요.')
                except:
                    print('Selection에 있는 "정수"로 다시 입력해주세요')



            # order.set_sugar_cnt(swt)

            os.system('cls')
            table = Table(title="========== Size ==========", show_header=True, header_style="bold magenta")
            table.add_column('Selction', width=10, justify='center')
            table.add_column('Size', width=30, justify='left')
            table.add_column('Extra Charge', width=10, justify='center')

            table.add_row('1', 'Tall', '+0원')
            table.add_row('2', 'Grande','+700원')
            table.add_row('3', 'Venti','+1000원')

            console = Console()
            console.print(table)
            size_list = ['Tall', 'Grande', 'Venti']


            while True:
                try:
                    size = int(input('음료 사이즈를 선택해주세요>>'))
                    if size == 2:
                        add_on += 700
                    elif size == 3:
                        add_on += 1000
                    if size in [1, 2, 3]:
                        break
                    else:
                        print('다시 입력해주세요.')
                except:
                    print('Selection에 있는 "정수"로 다시 입력해주세요')



            bags_str = f'{temp_list[temp - 1]} {coffee_menu_list[menu_name][0]} {qntt}개, 당도는 {swt_list[swt - 1]}, 사이즈는 {size_list[size - 1]}'
            print(f'{bags_str}를 선택하셨습니다 ')
            StarBucks_Repository.order_str.append(bags_str)

            order = StarBucks(coffee_menu_list[menu_name][0])
            # order.set_name(coffee_menu_list[menu_name][0])
            order.set_temp(temp_list[temp - 1])
            order.set_quantity(qntt)
            order.set_size(size_list[size - 1])
            order.set_sugar_cnt(swt_list[swt - 1])
            if temp == 2:
                order.set_amnt_ice(ice_list[ice - 1])
            order.set_price((coffee_menu_list[menu_name][1] + add_on) * qntt)
            # print(order)

        elif menu_name == 0:
            return

        else:
            print('잘못 선택하셨습니다. 다시 선택해주세요')
            time.sleep(1)
            return self.coffee_menu()

        # StarBucks_Repository.sales_statement.append(order)

        input()
        return order


    def blended_menu(self):
            blended_menu_list = {
                1: ('녹차 프라투치노', 6500),
                2: ('자바칩 프라푸치노', 7000),
                3: ('피치 아사이 리프레셔', 6000),
                4: ('딸기 아사이 리프레셔', 6000),
                5: ('망고 리프레셔', 6000)
            }
            blended_menu_str = """
    ---------Blended Menu--------
    1. 녹차 프라푸치노          =========     6500 원
    2. 자바칩 프라푸치노        =========     7000 원
    3. 피치 아사이 리프레셔     =========     6000 원
    4. 딸기 아사이 리프레셔     =========     6000 원
    5. 망고 리프레셔           =========     6000 원
    0. 이전으로 돌아가기
    ----------------------------          
    선택:>>"""

            re_flag = 1

            os.system('cls')
            table = Table(title="========== Coffee Menu ==========", show_header=True, header_style="bold magenta")
            table.add_column('Selction', width=10, justify='center')
            table.add_column('Menu', width=30, justify='left')
            table.add_column('Price', width=10, justify='center')

            table.add_row('1', '녹차 프라푸치노', '6500원')
            table.add_row('2', '자바칩 프라푸치노', '7000원')
            table.add_row('3', '피치 아사이 리프레셔', '6000원')
            table.add_row('4', '딸기 아사이 리프레셔', '6000원')
            table.add_row('5', '망고 리프레셔', '6000원')
            table.add_row('0', '이전으로 돌아가기', '')

            console = Console()
            console.print(table)

            while True:
                try:
                    menu_name = int(input('메뉴를 선택해 주세요.>>'))
                    break
                except:
                    print('"정수"로 다시 입력해주세요')


            add_on = 0

            if menu_name in blended_menu_list:
                print(f'{blended_menu_list[menu_name][0]}를 선택하셨습니다.')

                while True:
                    try:
                        qntt = int(input('\n수량을 입력해 주세요>>'))
                        break
                    except:
                        print('Selection에 있는 "정수"로 다시 입력해주세요')


                os.system('cls')
                table = Table(title="========== Sweetness ==========", show_header=True, header_style="bold magenta")
                table.add_column('Selction', width=10, justify='center')
                table.add_column('Sweetness', width=30, justify='left')
                table.add_column('Extra Charge', width=10, justify='center')

                table.add_row('1', '더 달게 120%', '+300원')
                table.add_row('2', '보통 100%', '+0원')
                table.add_row('3', '덜 달게 80%', '+0원')

                console = Console()
                console.print(table)

                swt_list = ['120%', '100%', '80%']
                while True:
                    try:
                        swt = int(input('당도를 선택해주세요>>'))
                        if swt == 1:
                            add_on += 300
                        if swt in [1,2,3]:
                                break
                        else:
                            print('Selection에 있는 "정수"로 다시 입력해주세요')
                    except:
                        print('Selection에 있는 "정수"로 다시 입력해주세요')







                os.system('cls')
                table = Table(title="========== Size ==========", show_header=True, header_style="bold magenta")
                table.add_column('Selction', width=10, justify='center')
                table.add_column('Size', width=30, justify='left')
                table.add_column('Extra Charge', width=10, justify='center')

                table.add_row('1', 'Tall', '+0원')
                table.add_row('2', 'Grande', '+700원')
                table.add_row('3', 'Venti', '+1000원')

                console = Console()
                console.print(table)
                size_list = ['Tall', 'Grande', 'Venti']

                while True:
                    try:
                        size = int(input('음료 사이즈를 선택해주세요>>'))
                        if size == 2:
                            add_on += 700
                        elif size == 3:
                            add_on += 1000
                        if size in [1, 2, 3]:
                            break
                    except:
                        print('Selection에 있는 "정수"로 다시 입력해주세요')


                os.system('cls')
                table = Table(title="========== Ice ==========", show_header=True, header_style="bold magenta")
                table.add_column('Selction', width=10, justify='center')
                table.add_column('Ice amount', width=30, justify='left')
                table.add_column('Extra Charge', width=10, justify='center')

                table.add_row('1', 'Extra', '+0원')
                table.add_row('2', 'Normal', '+0원')
                table.add_row('3', 'Less', '+0원')

                console = Console()
                console.print(table)

                ice_list = ['Extra', 'Normal', 'Less']

                while True:
                    try:
                        ice = int(input('얼음양을 선택해주세요>>'))
                        if ice in [1,2,3]:
                            break
                    except:
                        print('Selection에 있는 "정수"로 다시 입력해주세요')

            elif menu_name == 0:
                return
            else:
                print('잘못 선택하셨습니다. 다시 선택해주세요')
                time.sleep(1)
                return self.blended_menu()

            bags_str = f'{blended_menu_list[menu_name][0]} {qntt}개, 당도는 {swt_list[swt - 1]}, 사이즈는 {size_list[size - 1]}'
            print(f'{bags_str}를 선택하셨습니다 ')
            StarBucks_Repository.order_str.append(bags_str)

            order = StarBucks(blended_menu_list[menu_name][0])
            # order.set_name(coffee_menu_list[menu_name][0])
            order.set_quantity(qntt)
            order.set_size(size_list[size - 1])
            temp_list = ['Hot', 'Iced']
            order.set_temp(temp_list[1])
            order.set_amnt_ice(ice_list[ice - 1])
            order.set_sugar_cnt(swt_list[swt - 1])
            order.set_price((blended_menu_list[menu_name][1] + add_on) * qntt)

            # print(order)

            # bags = FinalPage.add_shopping_bag(999,order)
            # print("장바구니에 추가 되었습니다.")
            # print(bags)

            # StarBucks_Repository.sales_statement.append(order)

            return order


            # return FinalPage.cart_price_sum(999,bags)

    def noncoffee_menu(self):
        noncoffee_menu_list = {
            1: ('녹차 라테', 6500),
            2: ('라이트 자몽 피지오', 5800),
            3: ('복숭아 아이스 티', 5900),
            4: ('딸기에이드', 5500),
            5: ('레몬에이드', 5500)
        }
    #     blended_menu_str = """
    # ---------Blended Menu--------
    # 1. 녹차 라테             =========     6500 원
    # 2. 라이트 자몽 피지오     =========     5800 원
    # 3. 복숭아 아이스 티      =========     5900 원
    # 4. 딸기에이드           =========     5500 원
    # 5. 레몬에이드           =========     5500 원
    # 0. 이전으로 돌아가기
    # ----------------------------
    # 선택:>>"""

        re_flag = 1

        os.system('cls')
        table = Table(title="========== Non Coffee Menu ==========", show_header=True, header_style="bold magenta")
        table.add_column('Selction', width=10, justify='center')
        table.add_column('Menu', width=30, justify='left')
        table.add_column('Price', width=10, justify='center')

        table.add_row('1', '녹차 라테', '6500원')
        table.add_row('2', '라이트 자몽 피지오', '5800원')
        table.add_row('3', '복숭아 아이스 티', '5900원')
        table.add_row('4', '딸기에이드', '5500원')
        table.add_row('5', '레몬에이드', '5500원')
        table.add_row('0', '이전으로 돌아가기', '')

        console = Console()
        console.print(table)
        while True:
            try:
                menu_name = int(input('메뉴를 선택해 주세요.>>'))
                if menu_name in [1,2,3,4,5,0]:
                    break
            except:
                print('Selection에 있는 "정수"로 다시 입력해주세요')


        add_on = 0

        if menu_name in noncoffee_menu_list:
            print(f'{noncoffee_menu_list[menu_name][0]}를 선택하셨습니다.')

            temp_list = ['Hot', 'Iced']

            os.system('cls')
            table = Table(title="========== Hot / Cold ==========", show_header=True, header_style="bold magenta")
            table.add_column('Selction', width=10, justify='center')
            table.add_column('Menu', width=30, justify='left')
            table.add_column('Menu', width=10, justify='center')

            table.add_row('1', 'Hot', '+0원')
            table.add_row('2', 'Iced', '+500원')

            console = Console()
            console.print(table)

            while True:
                try:
                    temp = int(input('원하시는 음료 종류를 선택해 주세요 >>'))
                    if temp in [1,2]:
                        if temp == 2:
                            add_on += 500
                            ice_list = ['Extra', 'Normal', 'Less']

                            os.system('cls')
                            table = Table(title="========== Ice ==========", show_header=True, header_style="bold magenta")
                            table.add_column('Selction', width=10, justify='center')
                            table.add_column('Ice amount', width=30, justify='left')
                            table.add_column('Extra Charge', width=10, justify='center')

                            table.add_row('1', 'Extra', '+0원')
                            table.add_row('2', 'Normal', '+0원')
                            table.add_row('3', 'Less', '+0원')

                            console = Console()
                            console.print(table)

                            while True:
                                try:
                                    ice = int(input('얼음양을 선택해주세요>>'))
                                    if ice in [1,2,3]:
                                        break
                                except:
                                    print('"정수"로 다시 입력해주세요')
                        break

                except:
                    print('Selection에 있는 "정수"로 다시 입력해주세요')


            while True:
                try:
                    qntt = int(input('수량을 입력해 주세요\n>>'))
                    break
                except:
                    print('"정수"로 다시 입력해주세요')

            swt_list = ['120%', '100%', '80%']

            os.system('cls')
            table = Table(title="========== Sweetness ==========", show_header=True, header_style="bold magenta")
            table.add_column('Selction', width=10, justify='center')
            table.add_column('Sweetness', width=30, justify='left')
            table.add_column('Extra Charge', width=10, justify='center')

            table.add_row('1', '더 달게 120%', '+300원')
            table.add_row('2', '보통 100%', '+0원')
            table.add_row('3', '덜 달게 80%', '+0원')

            console = Console()
            console.print(table)

            while True:
                try:
                    swt = int(input('당도를 선택해주세요>>'))

                    if swt == 1:
                        add_on += 300

                    if swt in [1,2,3]:
                        break
                except:
                    print('Selection에 있는 "정수"로 다시 입력해주세요')

            size_list = ['Tall', 'Grande', 'Venti']
            os.system('cls')
            table = Table(title="========== Size ==========", show_header=True, header_style="bold magenta")
            table.add_column('Selction', width=10, justify='center')
            table.add_column('Size', width=30, justify='left')
            table.add_column('Extra Charge', width=10, justify='center')

            table.add_row('1', 'Tall', '+0원')
            table.add_row('2', 'Grande', '+700원')
            table.add_row('3', 'Venti', '+1000원')

            console = Console()
            console.print(table)

            while True:
                try:
                    size = int(input('음료 사이즈를 선택해주세요>>'))

                    if size == 2:
                        add_on += 700
                    elif size == 3:
                        add_on += 1000

                    if size in [1,2,3]:
                        break
                except:
                    print('Selection에 있는 "정수"로 다시 입력해주세요')


        elif menu_name == 0:
            return
        else:
            print('잘못 선택하셨습니다. 다시 선택해주세요')
            time.sleep(1)
            return self.noncoffee_menu()

        bags_str = f'{noncoffee_menu_list[menu_name][0]} {qntt}개, 당도는 {swt_list[swt - 1]}, 사이즈는 {size_list[size - 1]}'
        print(f'{bags_str}를 선택하셨습니다')
        StarBucks_Repository.order_str.append(bags_str)

        order = StarBucks(noncoffee_menu_list[menu_name][0])
        # order.set_name(coffee_menu_list[menu_name][0])
        order.set_quantity(qntt)
        order.set_temp(temp_list[temp-1])
        order.set_size(size_list[size - 1])
        order.set_sugar_cnt(swt_list[swt - 1])
        if temp == 2:
            order.set_amnt_ice(ice_list[ice - 1])
        order.set_price((noncoffee_menu_list[menu_name][1] + add_on) * qntt)

        # print(order)

        # bags = FinalPage.add_shopping_bag(999,order)
        # print("장바구니에 추가 되었습니다.")
        # print(bags)

        # StarBucks_Repository.sales_statement.append(order)
        return order




# class Stat:
#     def __init__(self,ave_sales):
#         self.ave_sales = ave_sales
#
#     def cal_sales(self, orders):
#         for item in orders:
#
#
#
#
#


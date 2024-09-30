from stbk_entity import StarBucks
from stbk_entity import Coffee
from stbk_entity import NonCoffee
import pickle
import openpyxl as op

"""
- 커피 : 온도(Hot/Iced), 사이즈, 얼음량, 당도, 컵(텀블러 유무), 수량, 디카페인, [상품명, 샷(연하게, 보통, 진하게(샷추가+500) ), 가격 ]
- 음료 : 온도(Hot/Iced), 사이즈, 얼음량, 당도, 컵(텀블러 유무), 수량, [상품명, 휘핑 유무, 가격]

"""

class StarBucks_Repository:

    def __init__(self):
        print('StarBucks_Repository 인스턴스가 생성되었습니다.')
        self.stbks = [
            Coffee('아메리카노','Hot','tall', '보통','보통','매장','False','2','3'),
            Coffee('아메리카노','Iced','grande', '보통','보통','take-out','False','3','2'),
            Coffee('카라멜마끼아또','Hot','grande', '보통','보통','take-out','True','2','1'),
            NonCoffee('자바칩 프라푸치노','Iced','grande', '보통','보통','매장','많이','2'),
            NonCoffee('자몽허니블래티','Iced','grande', '적게','더 달게','take-out','없이','5'),
            NonCoffee('레몬에이드','Iced','Venti', '보통','보통','take-out','없이','1'),
            Coffee('카페모카','Hot','tall', '없이','보통','매장','False','2','3')
        ]

    def find_all(self):
            return self.stbks


        # TODO 추가했으면 하는것..커피/음료에서 얼음이 안들어가는것(HOT, 프라푸치노,...),휘핑이 원래 안들어가는것들(아메리카노, 자허블, ...)은 옵션이 안보이게 하고 싶다...
    def push(self, stbks):
        return self.create_workbook(stbks)

    def create_workbook(self, stbks):
        wb = op.Workbook()
        ws = wb.active
        ws['A1'].value = '번호'
        ws['B1'].value = '상품명'
        ws['C1'].value = '수량'
        ws['D1'].value = '매출액'
        # ws['E1'].value = '마진'
        # https://post.naver.com/viewer/postView.nhn?volumeNo=17083716&memberNo=6347519

        no = 1
        for stbk in stbks:
            ws[f'A{no + 1}'] = stbk.get_id()
            ws[f'B{no + 1}'] = stbk.get_name()
            ws[f'C{no + 1}'] = stbk.get_quantity()
            ws[f'D{no + 1}'] = stbk.get_sv()
            no += 1

        wb.save('starbucks.xlsx')
        print('☕☕엑셀파일이 생성되었습니다.☕☕')
        return wb




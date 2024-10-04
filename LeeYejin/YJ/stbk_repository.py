
from stbk_entity import StarBucks,FinalPage
import pickle
import openpyxl as op
import pandas as pd
from datetime import datetime


"""
- 커피 : 온도(Hot/Iced), 사이즈, 얼음량, 당도, 컵(텀블러 유무), 수량, 디카페인, [상품명, 샷(연하게, 보통, 진하게(샷추가+500) ), 가격 ]
- 음료 : 온도(Hot/Iced), 사이즈, 얼음량, 당도, 컵(텀블러 유무), 수량, [상품명, 휘핑 유무, 가격]
"""


class StarBucks_Repository:
    def __init__(self):
        print('StarBucksRepository 인스턴스가 생성되었습니다.')

        try:
            with open('stbks.pkl', 'rb') as f:
                self.stbks = pickle.load(f)

                if len(self.stbks) > 0:
                    self.stbks = self.stbks.append()

                print('> stbks.pickle이 로드되었습니다.')
        except FileNotFoundError:
            # 파일이 존재하지 않는 경우
            self.stbks = []

    def __del__(self):
        """
        소멸자 메소드 : 인스턴스가 메모리에서 해제될때 호출되는 메소드
        """
        with open('stbks.pkl', 'wb') as f:
            pickle.dump(self.stbks, f)
            print('❤️❤️매출내역을 성공적으로 저장했습니다. 다음에 만나요❤️❤️')

    # print('StarBucks_Repository 인스턴스가 생성되었습니다.')
        # self.stbks = [
        #     Coffee('아메리카노','Hot','tall', '보통','보통','매장','False','2','3'),
        #     Coffee('아메리카노','Iced','grande', '보통','보통','take-out','False','3','2'),
        #     Coffee('카라멜마끼아또','Hot','grande', '보통','보통','take-out','True','2','1'),
        #     NonCoffee('자바칩 프라푸치노','Iced','grande', '보통','보통','매장','많이','2'),
        #     NonCoffee('자몽허니블래티','Iced','grande', '적게','더 달게','take-out','없이','5'),
        #     NonCoffee('레몬에이드','Iced','Venti', '보통','보통','take-out','없이','1'),
        #     Coffee('카페모카','Hot','tall', '없이','보통','매장','False','2','3')
        # ]

    def find_all(self):
        return self.stbks

    # TODO 추가했으면 하는것..커피/음료에서 얼음이 안들어가는것(HOT, 프라푸치노,...),휘핑이 원래 안들어가는것들(아메리카노, 자허블, ...)은 옵션이 안보이게 하고 싶다...
    def push(self, stbks):
        return self.create_workbook(stbks)

    def create_workbook(self, stbks):
        wb = op.Workbook()
        ws = wb.active
        ws['A1'].value = '주문일시'
        ws['B1'].value = '번호'
        ws['C1'].value = '상품명'
        ws['D1'].value = 'HOT/ICED'
        ws['E1'].value = '사이즈'
        ws['F1'].value = '얼음량'
        ws['G1'].value = '당도'
        ws['H1'].value = '컵'
        ws['I1'].value = '수량'
        ws['J1'].value = '매출액'

        no = 1
        now = datetime.now()
        for stbks in stbks:
            ws[f'A{no + 1}'] = now.strftime('%Y-%m-%d %H:%M:%S')
            ws[f'B{no + 1}'] = stbks.get_id()
            ws[f'C{no + 1}'] = stbks.get_name()
            ws[f'D{no + 1}'] = stbks.get_temp()
            ws[f'E{no + 1}'] = stbks.get_size()
            ws[f'F{no + 1}'] = stbks.get_amnt_ice()
            ws[f'G{no + 1}'] = stbks.get_sugar_cnt()
            ws[f'H{no + 1}'] = stbks.get_cup()
            ws[f'I{no + 1}'] = stbks.get_quantity()
            ws[f'J{no + 1}'] = stbks.get_price()
            no += 1
        wb.save('stbks.xlsx')
        print('☕☕엑셀파일이 생성되었습니다.☕☕')
        return wb

    def coffee_menu(self):
        add_on = 0
        coffee_menu_list = {
            1: ('아메리카노', 4500),
            2: ('카페라떼', 5000),
            3: ('돌체라떼', 5900),
            4: ('카페모카', 5500),
            5: ('캬라멜마끼아또', 5900)
        }
        coffee_menu_str = """
    ---------Coffee Menu--------
    1. 아메리카노 ======= 4500원
    2. 카페라떼 ========= 5000원
    3. 돌체라떼 ========= 5500원
    4. 바닐라 라떼 ====== 5500원
    5. 카라멜마키아토 ==== 5500원
    0. 이전으로 돌아가기
    ----------------------------          
    선택:  """

        menu_name = int(input(coffee_menu_str))

        if menu_name in coffee_menu_list:
            print(f'{coffee_menu_list[menu_name][0]}를 선택하셨습니다.')
            temp_list = ['HOT', 'ICED']
            temp = int(input("""
    ----HOT / ICED----
    1. HOT
    2. ICED (+500 원)    
    선택:    """))
            if temp == 2:
                add_on += 500

            qntt = int(input('\n수량을 입력해주세요\n'))

            swt_list = ['120%', '100%', '80%']
            swt = int(input('''
    ----당도----
    1. 더 달게(120% + 300원)
    2. 보통(100%)
    3. 덜 달게(80%)
    선택: '''))
            if swt == 1:
                add_on += 300

            # order.set_sugar_cnt(swt)
            size_list = ['Tall', 'Grande', 'Venti']
            size = int(input('''
    ----사이즈----
    1. Tall (+0 원)
    2. Grande (+700 원)
    3. Venti (+1000 원)
    선택: '''))
            if size == 2:
                add_on += 700
            elif size == 3:
                add_on += 1000

            print(
                f'{coffee_menu_list[menu_name][0]} {temp_list[temp - 1]} {qntt}개, 당도는 {swt_list[swt - 1]}, 사이즈는 {size_list[size - 1]}를 선택하셨습니다 ')

            order = StarBucks(coffee_menu_list[menu_name][0])
            # order.set_name(coffee_menu_list[menu_name][0])
            order.set_temp(temp)
            order.set_quantity(qntt)
            order.set_size(size_list[size - 1])
            order.set_sugar_cnt(swt)
            order.set_price((coffee_menu_list[menu_name][1] + add_on) * qntt)
            print(order)

        elif menu_name == 0:
            return

        else:
            print('잘못 선택하셨습니다. 다시 선택해주세요')
            return self.coffee_menu()

        return order

    def blended_menu(self):
        blended_menu_list = {
            1: ('녹차 프라투치노', 6500),
            2: ('자바칩 프라푸치노', 7000),
            3: ('피치 아사이 리프레셔', 6000),
            4: ('딸기 아사이 리프레셔', 6000),
            5: ('망고 리프레셔', 6000)
        }
        blended_menu_str = """
    ---------Blended Menu--------
    1. 녹차 프라푸치노          =========     6500 원
    2. 자바칩 프라푸치노        =========     7000 원
    3. 피치 아사이 리프레셔     =========     6000 원
    4. 딸기 아사이 리프레셔     =========     6000 원
    5. 망고 리프레셔           =========     6000 원
    0. 이전으로 돌아가기
    ----------------------------          
    선택:>>"""

        re_flag = 1
        menu_name = int(input(blended_menu_str))
        add_on = 0

        if menu_name in blended_menu_list:
            print(f'{blended_menu_list[menu_name][0]}를 선택하셨습니다.')

            qntt = int(input('\t수량을 입력해 주세요\n>>'))

            swt_list = ['120%', '100%', '80%']

            swt = int(input('''

    ---- 당도를 입력해주세요 ---- 
    1 = 더 달게 120% (+300원)
    2 = 보통 100%
    3 = 덜 달게 80%
    >>'''))

            if swt == 1:
                add_on += 300

            size_list = ['Tall', 'Grande', 'Venti']
            size = int(input(''' 
    ---- 사이즈를 입력해주세요. ----
    1 = Tall(+0원)
    2 = Grande(+700원)
    3 = Venti(+1000원)
    >>'''))
            if size == 2:
                add_on += 700
            elif size == 3:
                add_on += 1000

            ice_list = ['Extra', 'Normal', 'Less']
            ice = int(input('''
    ---- 얼음양을 입력해주세요. ----
    1 = Extra
    2 = Normal
    3 = Less
    >>'''))
        elif menu_name == 0:
            return
        else:
            print('잘못 선택하셨습니다. 다시 선택해주세요')
            return self.blended_menu()

        print(
            f'{blended_menu_list[menu_name][0]} {qntt}개, 당도는 {swt_list[swt - 1]}, 사이즈는 {size_list[size - 1]}를 선택하셨습니다 ')

        order = StarBucks(blended_menu_list[menu_name][0])
        # order.set_name(coffee_menu_list[menu_name][0])
        order.set_quantity(qntt)
        order.set_size(size_list[size - 1])
        order.set_amnt_ice(ice_list[ice - 1])
        order.set_sugar_cnt(swt_list[swt - 1])
        order.set_price((blended_menu_list[menu_name][1] + add_on) * qntt)

        # print(order)

        # bags = FinalPage.add_shopping_bag(999,order)
        # print("장바구니에 추가 되었습니다.")
        # print(bags)

        return order

        # return FinalPage.cart_price_sum(999,bags)

    def noncoffee_menu(self):
        blended_menu_list = {
            1: ('녹차 라테', 6500),
            2: ('라이트 자몽 피지오', 5800),
            3: ('복숭아 아이스 티', 5900),
            4: ('딸기에이드', 5500),
            5: ('레몬에이드', 5500)
        }
        blended_menu_str = """
    ---------Blended Menu--------
    1. 녹차 라테             =========     6500 원
    2. 라이트 자몽 피지오     =========     5800 원
    3. 복숭아 아이스 티      =========     5900 원
    4. 딸기에이드           =========     5500 원
    5. 레몬에이드           =========     5500 원
    0. 이전으로 돌아가기
    ----------------------------          
    선택:>>"""

        re_flag = 1
        menu_name = int(input(blended_menu_str))
        add_on = 0

        if menu_name in blended_menu_list:
            print(f'{blended_menu_list[menu_name][0]}를 선택하셨습니다.')

            qntt = int(input('수량을 입력해 주세요\n>>'))

            swt_list = ['120%', '100%', '80%']

            swt = int(input('''

    ---- 당도를 입력해주세요 ---- 
    1=120% (+300)
    2=100%
    3=80%
    >>'''))

            if swt == 1:
                add_on += 300

            size_list = ['Tall', 'Grande', 'Venti']
            size = int(input(''' 
    ---- 사이즈를 입력해주세요. ----
    1 = Tall(+0)
    2 = Grande(+700)
    3 = Venti(+1000)
    >>'''))
            if size == 2:
                add_on += 700
            elif size == 3:
                add_on += 1000

            ice_list = ['Extra', 'Normal', 'Less']
            ice = int(input('''
    ---- 얼음양을 입력해주세요. ----
    1 = Extra
    2 = Normal
    3 = Less
    >>'''))
        elif menu_name == 0:
            return
        else:
            print('잘못 선택하셨습니다. 다시 선택해주세요')
            return self.blended_menu()

        print(
            f'{blended_menu_list[menu_name][0]} {qntt}개, 당도는 {swt_list[swt - 1]}, 사이즈는 {size_list[size - 1]}를 선택하셨습니다 ')

        order = StarBucks(blended_menu_list[menu_name][0])
        # order.set_name(coffee_menu_list[menu_name][0])
        order.set_quantity(qntt)
        order.set_size(size_list[size - 1])
        order.set_amnt_ice(ice_list[ice - 1])
        order.set_sugar_cnt(swt_list[swt - 1])
        order.set_price((blended_menu_list[menu_name][1] + add_on) * qntt)

        # print(order)

        # bags = FinalPage.add_shopping_bag(999,order)
        # print("장바구니에 추가 되었습니다.")
        # print(bags)

        return order


class FinalPage:


    def __init__(self):
        pass

    # def shopping_bag_display(self):




    def cart_price_sum(self,bags):
        sum_price = 0
        for order_price in bags:
            sum_price += order_price.get_price()
            # print(sum_price)

        # print(sum_price)
        return sum_price

    def carrier_sel(self):
        print('사용하고 계신 통신사를 선택해 주세요.')
        str_carrier_menu = '''
    1. SKT (10% 할인)
    2. KT   (10% 할인)
    3. LGU (5% 할인)
    4. 해당없음.
    '''
        input_carrier = input(str_carrier_menu)
        return input_carrier

    skt = 0.1
    kt = 0.1
    lgu = 0.05

    order_nums = 0

    def calculation(self, carrier, total_price): # 통신사 없을 시 추가 해야함

        match carrier:
            case '1':
                discount_price = total_price * (1-FinalPage.skt)
            case '2':
                discount_price = total_price * (1-FinalPage.kt)
            case '3':
                discount_price = total_price * (1-FinalPage.lgu)
            case '4':
                print('해당되는 할인 항목이 없습니다.')
                discount_price = total_price
            case _:
                print('잘못된 통신사입니다.')
                return 0

        return discount_price

    def payment(self):      # 1) 결제가능 제한시간, 2) 일정시간(3초) 후 결제완료 메세지 표출
        method = input('결제 수단을 선택하세요\n1)신용카드 ,2)현금결제, 3)각종 pay\n')
        flag = 1

        match method:
            case '1':   # 신용카드
                print('카드를 칩 리더기에 넣어주세요')
                for i in track(range(3)):
                    time.sleep(0.5)

            case '2':   # 현금결제
                print('카운터에서 결제 해주세요')
                time.sleep(0.5)
                print('감사합니다!')
            case '3':   # 삼성, 애플 pay
                print('핸드폰을 테그에 접촉해주세요')
                for i in track(range(3)):
                    time.sleep(0.5)
            case _:
                print('잘못된 선택입니다.')
                flag = 0

        if flag != '0':
            print('Requesting transaction....')
            time.sleep(0.5)
            print('Waiting for response....')
            time.sleep(0.2)
            print('Data transmit successful....')
            time.sleep(0.5)
            print('Receiving encrypted transaction data....')
            time.sleep(0.1)
            print('Transaction was successful....')
            time.sleep(0.2)

            print('성장적으로 결제 되었습니다.')
        return flag

    members = {'20745440': 2}

    def membership(self):
            member = input('\t멤버쉽 있으신가요?? 1. yes, 2. no\n>>')
            flag = 1

            match member:
                case '1' :
                    # cell_number = input('핸드폰 번호를 입력해주세요')        # try except 가능
                    FinalPage.add_point(999)
                case '2':
                    mem_join = input('\t멤버쉽에 가입하시겠습니까? [1=y]/2=n \n>>')
                    if mem_join != '2':
                        FinalPage.join_member(999)
                    else :
                        print('\t결제화면으로 이동합니다.')
                case _:
                    print('잘못 누르셨습니다.')
                    flag = 0

            return flag

    def add_point(self):
        cell_number = input('핸드폰 번호를 입력해주세요\n>>')
        try :
            FinalPage.members[cell_number] += 1 # members : 멤버쉽 딕셔너리들, 값들은 숫자여야 함.
            print(f'1포인트 적립이 완료 되었습니다. {cell_number}님의 현재 포인트는 {FinalPage.members[cell_number]} 포인트 입니다.')
        except:
            print('등록되지 않은 회원입니다. 회원가입을 진행합니다.')
            FinalPage.join_member(999)
            # Final_page.members[cell_number] = 1
            # print(f'1포인트 적립이 완료되었습니다. {cell_number}님의 현재 포인트는 {Final_page.members[cell_number]} 포인트 입니다.')


    def join_member(self):
        cell_number = input('핸드폰 번호를 입력해주세요\n>>')
        if cell_number in FinalPage.members:
            FinalPage.members[cell_number] += 1  # members : 멤버쉽 딕셔너리들, 값들은 숫자여야 함.
            print('가입된 회원 정보가 있습니다.')
            print(f'1포인트 적립이 완료 되었습니다. {cell_number}님의 현재 포인트는 {FinalPage.members[cell_number]} 포인트 입니다.')
        else:
            FinalPage.members[cell_number] = 1    # 딕셔너리에 멤버 추가
            print(f'회원가입이 완료되었습니다. {cell_number}님의 현재 포인트는 {FinalPage.members[cell_number]} 포인트 입니다.')

    def receipt(self):
        print('주문이 완료 되었습니다. 감사합니다.')
        FinalPage.order_nums += 1
        print(f'주문번호는 {FinalPage.order_nums}번 입니다. ')


class Welcome:
    def __init__(self):
        pass
    def welcome_page(self):
        print('''
=================================================================================
__        __         _                                         _               
\ \      / /   ___  | |   ___    ___    _ __ ___     ___      | |_    ___      
 \ \ /\ / /   / _ \ | |  / __|  / _ \  | '_ ` _ \   / _ \     | __|  / _ \     
  \ V  V /   |  __/ | | | (__  | (_) | | | | | | | |  __/     | |_  | (_) |    
   \_/\_/     \___| |_|  \___|  \___/  |_| |_| |_|  \___|      \__|  \___/                                                
 ____                   _                    _                      _          
/ ___|    ___    ___   | |_    __ _   _ __  | |__    _   _    ___  | | __  ___ 
\___ \   / _ \  / _ \  | __|  / _` | | '__| | '_ \  | | | |  / __| | |/ / / __|
 ___) | |  __/ | (_) | | |_  | (_| | | |    | |_) | | |_| | | (__  |   <  \__ \\
|____/   \___|  \___/   \__|  \__,_| |_|    |_.__/   \__,_|  \___| |_|\_\ |___/


================================================================================

''')
        input('계속하려면 엔터를 눌러주세요...')

    def where_menu(self):
        togo_str = '''
    --------- 어디에서 드시나요? ---------
    1. Dine-in (매장)
    2. To-Go (포장)
    0. 직원 호출
    -----------------------------------       
    선택 >> '''
        where = input(togo_str)
        # 관리자 모드 실행
        if where == '-1':
            password = input('>> 관리자 모드를 실행합니다.\n비밀번호를 입력하세요.\n비밀번호 :')
            manager_password = '1234'
            if password == manager_password:
                self.manager_mode()
            else:
                print('>> 비밀번호를 틀렸습니다.')

    def manager_mode(self):
        manager_mode = """
-----------매니저 관리 모드------------
1. 매출 정보 조회
2. 매출 정보 엑셀내보내기
0. 종료  
-------------------------------------
선택:                      """
        while True:
            manager_choice = input(manager_mode)
            match manager_choice:
                # 매출 정보 조회
                case '1':
                    stbks = self.find_all()
                    self.print_stbks(stbks)
                    print('😀😀매출정보를 불러왔습니다.😀😀')
                # 매출 액셀로 내보내기
                case '2':
                    stbks = self.stbk_service.find_all()
                    self.stbk_service.push(stbks)
                case '0':
                    return
                case _:
                    print('>> 잘못 선택 하셨습니다.')


    def byebye(self):
        str_bye = '''
                                    (  )   (   )  )
                                 ) (   )  (  (
                                 ( )  (    ) )
                                 _____________
                                <_____________> ___
                                |             |/ _ \\
                                |               | | |
                                |               |_| |
                             ___|             |\___/
                            /    \___________/    \\
                            \_____________________/    
        '''

        print('☕☕ 서-타-벅스를 이용해 주셔서 감사합니다. ☕☕')
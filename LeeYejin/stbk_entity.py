from stbk_entity import StarBucks
from datetime import datetime
import openpyxl as op

"""
- 커피 : 온도(Hot/Iced), 사이즈, 얼음량, 당도, 컵(텀블러 유무), 수량, 디카페인, [상품명, 샷(연하게, 보통, 진하게(샷추가+500) ), 가격 ]
- 음료 : 온도(Hot/Iced), 사이즈, 얼음량, 당도, 컵(텀블러 유무), 수량, [상품명, 휘핑 유무, 가격]
"""

class StarBucks_Repository:
    def __init__(self):
        print('StarBucks_Repository 인스턴스가 생성되었습니다.')

        self.stbks = [
            StarBucks('아메리카노','Hot','tall', '보통','보통','매장','3', '5000'),
            StarBucks('아메리카노','Iced','grande', '보통','보통','take-out','3','5500'),
            StarBucks('카라멜마끼아또','Hot','grande', '보통','보통','take-out','2','6500'),
            ]


#
#     def coffee_menu(self):
#         coffee_menu_list = {
#             1: ('아메리카노', 4500),
#             2: ('카페라떼', 5000),
#             3: ('돌체라떼', 5900),
#             4: ('카페모카', 5500),
#             5: ('캬라멜마끼아또', 5900)
#         }
#         coffee_menu_str = """
# ---------Coffee Menu--------
# 1. 아메리카노 ======= 4500원
# 2. 카페라떼 ========= 5000원
# 3. 돌체라떼 ========= 5500원
# 4. 바닐라 라떼 ====== 5500원
# 5. 카라멜마키아토 ==== 5500원
# 0. 이전으로 돌아가기
# ----------------------------
# 선택:                    """
#
#         re_flag = 1
#
#         while True:
#             menu_name = int(input(coffee_menu_str))
#
#             add_on = 0
#
#             if menu_name in coffee_menu_list:
#                 print(f'{coffee_menu_list[menu_name][0]}를 선택하셨습니다.')
#                 temp_list = ['HOT', 'ICED']
#                 temp = int(input("""
# ----HOT / ICED----
# 1. HOT
# 2. ICED
# 선택:    """))
#                 if temp == 2:
#                     amnt_ice_list = ['적게', '보통', '많이']
#                     amnt_ice = int(input('''
# ----얼음량----
# 1. 적게
# 2. 보통
# 3. 많게
# 선택: '''))
#                     order = StarBucks(coffee_menu_list[menu_name][0])
#                     order.set_amnt_ice(amnt_ice_list[amnt_ice -1])
#                 else:
#                     pass
#                 qntt = int(input('수량을 입력해주세요'))
#                 swt_list = ['120%', '100%', '80%']
#                 swt = int(input('''
# ----당도----
# 1. 더 달게(120%) +300
# 2. 보통(100%)
# 3. 덜 달게(80%)
# 선택: '''))
#                 if swt == 1:
#                     add_on += 300
#                 # order.set_sugar_cnt(swt)
#                 size_list = ['Tall', 'Grande', 'Venti']
#                 size = int(input('''
# ----사이즈----
# 1. Tall
# 2. Grande +700
# 3. Venti +1000
# 선택: '''))
#                 if size == 2:
#                     add_on += 700
#                 elif size == 3:
#                     add_on += 1000
#
#                 print(
#                     f'{coffee_menu_list[menu_name][0]} {temp_list[temp - 1]} {qntt}개, 당도는 {swt_list[swt - 1]}, 사이즈는 {size_list[size - 1]}를 선택하셨습니다 ')
#
#
#                 order = StarBucks(coffee_menu_list[menu_name][0])
#                 order.set_temp(temp)
#                 order.set_quantity(qntt)
#                 order.set_size(size_list[size - 1])
#
#                 order.set_sugar_cnt(swt)
#                 order.set_price(coffee_menu_list[menu_name][1]*qntt)
#                 print(order)
#             elif menu_name == 0:
#                 return
#
#             else:
#                 print('잘못 선택하셨습니다. 다시 선택해주세요')
#
#
#             print("장바구니에 추가되었습니다.")
#             orders = Shopping_bag.add_shopping_bag(order)
#             print(orders)
#
#             re_flag = input('더 주문하시겠습니까? [Y = 1, N = 0] :  ')
#             if re_flag == '0':
#                 break
#
#         return self.cart_price_sum(orders)
#
#
#
#
#     def cart_price_sum(self,orders):
#         sum_price = 0
#         for order_price in orders:
#             sum_price += order_price.get_price()
#
#         return sum_price


    def find_all(self):
            return self.stbks

    def push(self, orders):
        return self.create_workbook(orders)

    def create_workbook(self, orders):
        wb = op.Workbook()
        ws = wb.active
        ws['A1'].value = '주문일시'
        ws['B1'].value = '번호'
        ws['C1'].value = '상품명'
        ws['D1'].value = 'HOT/ICED'
        ws['E1'].value = '사이즈'
        ws['F1'].value = '얼음량'
        ws['G1'].value = '당도'
        ws['H1'].value = '컵'
        ws['I1'].value = '수량'
        ws['J1'].value = '매출액'

        no = 1
        now = datetime.now()
        for order in orders:
            ws[f'A{no + 1}'] = now.strftime('%Y-%m-%d %H:%M:%S')
            ws[f'B{no + 1}'] = order.get_id()
            ws[f'C{no + 1}'] = order.get_name()
            ws[f'D{no + 1}'] = order.get_temp()
            ws[f'E{no + 1}'] = order.get_size()
            ws[f'F{no + 1}'] = order.get_amnt_ice()
            ws[f'G{no + 1}'] = order.get_sugar_cnt()
            ws[f'H{no + 1}'] = order.get_cup()
            ws[f'I{no + 1}'] = order.get_quantity()
            ws[f'J{no + 1}'] = order.get_price()
            no += 1
        wb.save('starbucks.xlsx')
        print('☕☕엑셀파일이 생성되었습니다.☕☕')
        return wb




